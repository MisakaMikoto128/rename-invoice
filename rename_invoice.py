# -*- coding: utf-8 -*-
"""
发票 PDF 重命名工具 —— 严格双重校验（小写数字 ↔ 中文大写）

用法:
  python rename_invoice.py <PDF文件 或 目录> [<PDF文件 或 目录> ...]
  无参数时, 扫描脚本所在目录

  --silent        静默模式 (右键菜单专用): 多次并发调用通过文件锁合并为一次处理,
                  不输出到控制台, 仅写入日志.
  --summary       与 --silent 配合: 处理完后弹出 Tk 汇总窗口 (默认不弹).
  --xlsx          处理完后在目标文件夹生成 "发票汇总_YYYYMMDD-HHMMSS.xlsx",
                  含发票号码/开票日期/销售方/金额, 末尾合计公式.

可靠性策略 (重命名):
  1. 提取中文大写金额 (玖拾捌圆零壹分) 转为数字
  2. 提取所有 ¥X.XX 值
  3. 必须存在某个 ¥ 值精确等于中文大写转换后的数字
  4. 任一校验失败 -> 拒绝重命名, 保留原文件名
"""
import sys
import os
import re
import time
import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

try:
    import fitz  # PyMuPDF
except ImportError:
    print('[ERROR] 缺少依赖 PyMuPDF. 请运行: pip install pymupdf')
    try:
        input('按回车键退出...')
    except Exception:
        pass
    sys.exit(1)


SCRIPT_DIR = Path(__file__).resolve().parent
LOG_PATH = SCRIPT_DIR / 'rename_invoice.log'
QUEUE_FILE = SCRIPT_DIR / '.queue.txt'
QUEUE_LOCK = SCRIPT_DIR / '.queue.lock'
LEADER_LOCK = SCRIPT_DIR / '.leader.lock'
DEBOUNCE_SECONDS = 0.25

CHINESE_DIGITS = {
    '零': 0, '〇': 0,
    '壹': 1, '贰': 2, '叁': 3, '肆': 4, '伍': 5,
    '陆': 6, '柒': 7, '捌': 8, '玖': 9,
}
SMALL_UNITS = {'拾': 10, '佰': 100, '仟': 1000, '十': 10, '百': 100, '千': 1000}
BIG_UNITS = {'万': 10000, '亿': 100000000}

CHINESE_AMOUNT_CHARS = set(CHINESE_DIGITS) | set(SMALL_UNITS) | set(BIG_UNITS) | set('圆元角分整')

ALREADY_PREFIXED_RE = re.compile(r'^\d+(\.\d{1,2})?元-')
PRICE_RE = re.compile(r'¥\s*(\d+(?:\.\d{1,2})?)')

INVOICE_NO_LABEL_RE = re.compile(r'发\s*票\s*号\s*码\s*[:：]\s*(\d{8,25})')
INVOICE_NO_STANDALONE_RE = re.compile(r'^\s*(\d{15,25})\s*$', re.MULTILINE)
INVOICE_DATE_RE = re.compile(r'(\d{4}年\d{1,2}月\d{1,2}日)')
NAME_PREFIX_RE = re.compile(r'^\s*名\s*称\s*[:：]\s*')
COMPANY_SUFFIX_RE = re.compile(
    r'(?:有限公司|股份有限公司|股份公司|集团公司|集团|个体工商户|事务所|合伙企业|工作室|商行|经营部|分公司|店)$'
)


# ---------------------------------------------------------------------------
# 中文大写金额解析
# ---------------------------------------------------------------------------

def parse_chinese_int(s: str) -> int:
    """壹仟贰佰叁拾肆 -> 1234"""
    total = 0
    section = 0
    digit = 0
    for ch in s:
        if ch in CHINESE_DIGITS:
            digit = CHINESE_DIGITS[ch]
        elif ch in SMALL_UNITS:
            unit = SMALL_UNITS[ch]
            multiplier = digit if digit > 0 else 1
            section += multiplier * unit
            digit = 0
        elif ch in BIG_UNITS:
            section += digit
            if section == 0:
                section = 1
            total += section * BIG_UNITS[ch]
            section = 0
            digit = 0
    section += digit
    total += section
    return total


def chinese_amount_to_decimal(s: str) -> float:
    """玖拾捌圆零壹分 -> 98.01"""
    s = s.strip()
    yuan_idx = -1
    for marker in ('圆', '元'):
        idx = s.find(marker)
        if idx != -1:
            yuan_idx = idx
            break
    if yuan_idx == -1:
        raise ValueError(f'中文大写金额缺少 圆/元: {s!r}')

    int_part = s[:yuan_idx]
    frac_part = s[yuan_idx + 1:]

    integer = parse_chinese_int(int_part)

    jiao = 0
    fen = 0
    last_digit = 0
    for ch in frac_part:
        if ch in CHINESE_DIGITS:
            last_digit = CHINESE_DIGITS[ch]
        elif ch == '角':
            jiao = last_digit
            last_digit = 0
        elif ch == '分':
            fen = last_digit
            last_digit = 0
        elif ch == '整':
            pass

    return round(integer + jiao / 10 + fen / 100, 2)


def find_chinese_amounts(text: str):
    """从文本中找出所有候选中文大写金额."""
    candidates = []
    buf = []
    for ch in text:
        if ch in CHINESE_AMOUNT_CHARS:
            buf.append(ch)
        else:
            if buf:
                candidates.append(''.join(buf))
                buf = []
    if buf:
        candidates.append(''.join(buf))
    return [c for c in candidates if ('圆' in c or '元' in c) and any(d in c for d in CHINESE_DIGITS)]


# ---------------------------------------------------------------------------
# PDF 提取: 文本 -> 字段
# ---------------------------------------------------------------------------

def _read_pdf(pdf_path: Path):
    """
    返回 (full_text, first_page_blocks, page_width, error).
    blocks 是首页的 (x0, y0, x1, y1, text, ...) 列表, 用于布局判断.
    """
    try:
        doc = fitz.open(str(pdf_path))
    except Exception as e:
        return '', [], 0.0, f'无法打开 PDF: {e}'
    full_text = ''
    blocks = []
    page_width = 0.0
    try:
        for i, page in enumerate(doc):
            full_text += page.get_text() + '\n'
            if i == 0:
                page_width = float(page.rect.width)
                try:
                    blocks = page.get_text('blocks')
                except Exception:
                    blocks = []
    finally:
        doc.close()
    if not full_text.strip():
        return '', blocks, page_width, 'PDF 无文字层 (可能是扫描件), 不支持'
    return full_text, blocks, page_width, None


def _validate_amount_from_text(text: str):
    """三层校验后返回 (amount_str, reason). 成功 amount_str 形如 '98.01'."""
    chinese_candidates = find_chinese_amounts(text)
    if not chinese_candidates:
        return None, '未找到中文大写金额'

    chinese_decimals = []
    for c in chinese_candidates:
        try:
            chinese_decimals.append((c, chinese_amount_to_decimal(c)))
        except Exception:
            continue
    if not chinese_decimals:
        return None, f'中文大写金额无法解析: {chinese_candidates}'

    price_matches = PRICE_RE.findall(text)
    if not price_matches:
        return None, '未找到 ¥ 价格标记'
    prices = [round(float(p), 2) for p in price_matches]

    for cn_str, cn_val in chinese_decimals:
        for p in prices:
            if abs(p - cn_val) < 0.005:
                if cn_val + 0.005 < max(prices):
                    return None, (
                        f'中文大写金额 {cn_val} ({cn_str}) 不是最大 ¥ 值, '
                        f'最大 ¥ 值是 {max(prices)}, 异常'
                    )
                return f'{cn_val:.2f}', None

    return None, (
        f'中文大写金额与 ¥ 值不匹配. '
        f'中文: {[(s, v) for s, v in chinese_decimals]}, ¥ 值: {prices}'
    )


def _extract_invoice_no(text: str):
    """
    发票号码: 优先匹配 "发票号码: NUMBER" 同行格式;
    没有就找首个独占一行的 15-25 位数字 (旧版发票布局).
    """
    m = INVOICE_NO_LABEL_RE.search(text)
    if m:
        return m.group(1)
    m2 = INVOICE_NO_STANDALONE_RE.search(text)
    return m2.group(1) if m2 else None


def _extract_invoice_date(text: str):
    m = INVOICE_DATE_RE.search(text)
    return m.group(1) if m else None


def _strip_name_prefix(line: str) -> str:
    return NAME_PREFIX_RE.sub('', line).strip()


def _extract_seller_name(blocks, page_width: float, text_fallback: str):
    """
    用首页坐标判断: 公司名水平中点 > 页面中线 -> 销售方;
    没找到右半边的公司名时, 退回到文本顺序的第二个公司名 (旧版"label在前/value在后"布局).
    """
    if blocks and page_width > 0:
        center = page_width / 2
        right_candidates = []  # (cy, name)
        all_candidates = []
        for b in blocks:
            try:
                x0, y0, x1, y1, blk_text = b[0], b[1], b[2], b[3], b[4]
            except (IndexError, TypeError):
                continue
            cx = (x0 + x1) / 2
            cy = (y0 + y1) / 2
            for line in str(blk_text).splitlines():
                clean = _strip_name_prefix(line.strip())
                if clean and COMPANY_SUFFIX_RE.search(clean):
                    all_candidates.append((cy, cx, clean))
                    if cx > center:
                        right_candidates.append((cy, clean))
        if right_candidates:
            right_candidates.sort()  # 上方优先
            return right_candidates[0][1]
        if all_candidates:
            all_candidates.sort()
            return all_candidates[0][2]

    # 退路: 纯文本顺序解析 (购方在前, 销方在后)
    candidates = []
    for line in text_fallback.splitlines():
        clean = _strip_name_prefix(line.strip())
        if clean and COMPANY_SUFFIX_RE.search(clean):
            candidates.append(clean)
    if len(candidates) >= 2:
        return candidates[1]
    if len(candidates) == 1:
        return candidates[0]
    return None


def extract_invoice_metadata(pdf_path: Path):
    """
    一次读 PDF, 返回所有可提取的字段.
    返回 dict: amount(str|None), amount_reason(str|None),
               invoice_no(str|None), date(str|None), seller(str|None)
    """
    text, blocks, page_width, err = _read_pdf(pdf_path)
    if err:
        return {
            'amount': None, 'amount_reason': err,
            'invoice_no': None, 'date': None, 'seller': None,
        }
    amount, reason = _validate_amount_from_text(text)
    return {
        'amount': amount,
        'amount_reason': reason,
        'invoice_no': _extract_invoice_no(text),
        'date': _extract_invoice_date(text),
        'seller': _extract_seller_name(blocks, page_width, text),
    }


# ---------------------------------------------------------------------------
# 重命名 + 文件操作
# ---------------------------------------------------------------------------

def safe_target_path(directory: Path, name: str) -> Path:
    """目标文件已存在时, 追加 (2), (3) ... 不覆盖."""
    target = directory / name
    if not target.exists():
        return target
    stem, ext = os.path.splitext(name)
    n = 2
    while True:
        candidate = directory / f'{stem} ({n}){ext}'
        if not candidate.exists():
            return candidate
        n += 1


def log_line(line: str):
    ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        with open(LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(f'[{ts}] {line}\n')
    except Exception:
        pass


def process_pdf(pdf_path: Path):
    """
    返回 (status, message, metadata, final_path).
      status   : 'renamed' | 'skipped' | 'failed'
      message  : 给用户看的简短说明
      metadata : extract_invoice_metadata 的 dict (始终给出, 即使 status=skipped)
      final_path: 处理后文件的实际路径 (renamed 后是新名字, 否则是原路径)
    """
    name = pdf_path.name

    if ALREADY_PREFIXED_RE.match(name):
        meta = extract_invoice_metadata(pdf_path)
        return 'skipped', '已有价格前缀, 跳过', meta, pdf_path

    meta = extract_invoice_metadata(pdf_path)
    if meta['amount'] is None:
        return 'failed', meta['amount_reason'], meta, pdf_path

    new_name = f'{meta["amount"]}元-{name}'
    target = safe_target_path(pdf_path.parent, new_name)
    try:
        pdf_path.rename(target)
    except Exception as e:
        return 'failed', f'重命名失败: {e}', meta, pdf_path

    log_line(f'OK  {name}  ->  {target.name}  (金额={meta["amount"]})')
    return 'renamed', target.name, meta, target


def collect_pdfs(target: Path):
    if target.is_file():
        if target.suffix.lower() == '.pdf':
            return [target]
        return []
    if target.is_dir():
        return sorted(target.glob('*.pdf'))
    return []


# ANSI color (Windows 10+ terminal supports it)
def c(text, color):
    codes = {'red': '31', 'green': '32', 'yellow': '33', 'cyan': '36', 'gray': '90'}
    return f'\033[{codes.get(color, "0")}m{text}\033[0m'


# ---------------------------------------------------------------------------
# Excel 汇总输出
# ---------------------------------------------------------------------------

XLSX_HEADERS = ['发票文件名称', '发票号码', '开票日期', '销售方名称',
                '备注名称', '淘宝单号', '金额']

# Excel 人民币货币格式 (¥ 符号 + 千分位 + 两位小数, 负数加红色)
RMB_CURRENCY_FORMAT = '"¥"#,##0.00;[Red]"¥"-#,##0.00'


def write_summary_xlsx(rows, output_path: Path):
    """
    rows: list of dict, each having keys:
      filename, invoice_no, date, seller, amount (float or str)
    output_path: 目标 .xlsx 路径
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log_line('FAIL  导出 Excel 失败: 缺少 openpyxl 依赖, 请运行 pip install openpyxl')
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = '发票汇总'

    ws.append(XLSX_HEADERS)

    for r in rows:
        amt = r.get('amount')
        try:
            amt = float(amt) if amt is not None else None
        except (TypeError, ValueError):
            amt = None
        ws.append([
            r.get('filename', ''),
            r.get('invoice_no', '') or '',
            r.get('date', '') or '',
            r.get('seller', '') or '',
            '',  # 备注名称: 留空, 用户手填
            '',  # 淘宝单号: 留空, 用户手填
            amt,
        ])

    n_data = len(rows)
    last_data_row = 1 + n_data  # row 1 是表头
    total_row = last_data_row + 1

    ws.cell(row=total_row, column=6, value='合计')
    if n_data > 0:
        ws.cell(row=total_row, column=7,
                value=f'=SUM(G2:G{last_data_row})')
    else:
        ws.cell(row=total_row, column=7, value=0)

    # 样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    thin = Side(border_style='thin', color='B0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    right = Alignment(horizontal='right', vertical='center')

    for col_idx in range(1, 8):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for r in range(2, last_data_row + 1):
        for col_idx in range(1, 8):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = border
            if col_idx == 7:
                cell.number_format = RMB_CURRENCY_FORMAT
                cell.alignment = right
            elif col_idx in (2, 3):
                cell.alignment = center
            else:
                cell.alignment = left

    for col_idx in range(1, 8):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.border = border
        cell.fill = total_fill
        cell.font = total_font
        if col_idx == 7:
            cell.number_format = RMB_CURRENCY_FORMAT
            cell.alignment = right
        elif col_idx == 6:
            cell.alignment = right
        else:
            cell.alignment = center

    # 列宽 (估算, 中文等宽)
    widths = {'A': 60, 'B': 24, 'C': 16, 'D': 32, 'E': 18, 'F': 18, 'G': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return True


def _xlsx_output_path(target_dir: Path) -> Path:
    ts = datetime.datetime.now().strftime('%Y%m%d-%H%M%S')
    return target_dir / f'发票汇总_{ts}.xlsx'


def _row_from_result(status, message, metadata, final_path):
    """处理结果转换成 xlsx 行 (failed 的 amount 缺失就跳过, 由调用方判断)."""
    return {
        'filename': final_path.name,
        'invoice_no': metadata.get('invoice_no'),
        'date': metadata.get('date'),
        'seller': metadata.get('seller'),
        'amount': metadata.get('amount'),
    }


def _resolve_xlsx_target_dir(raw_args, fallback_pdfs):
    """决定 xlsx 输出目录: 第一个文件夹参数 / 否则第一个 pdf 的父目录 / 否则 cwd."""
    for raw in raw_args:
        p = Path(raw)
        if p.is_dir():
            return p
    if fallback_pdfs:
        return fallback_pdfs[0].parent
    return Path(os.getcwd())


# ---------------------------------------------------------------------------
# 队列 + 文件锁: 让多次并发调用合并为一次处理
# ---------------------------------------------------------------------------

def _open_lockfile(path: Path):
    """Open / create a 1-byte lock file. Returns file handle (binary, r+b)."""
    if not path.exists():
        try:
            path.write_bytes(b'\0')
        except Exception:
            pass
    return open(path, 'r+b')


def _acquire_blocking(fh):
    import msvcrt
    fh.seek(0)
    deadline = time.time() + 5.0
    while True:
        try:
            msvcrt.locking(fh.fileno(), msvcrt.LK_NBLCK, 1)
            return True
        except OSError:
            if time.time() > deadline:
                return False
            time.sleep(0.02)


def _acquire_nonblocking(fh):
    import msvcrt
    fh.seek(0)
    try:
        msvcrt.locking(fh.fileno(), msvcrt.LK_NBLCK, 1)
        return True
    except OSError:
        return False


def _release(fh):
    import msvcrt
    fh.seek(0)
    try:
        msvcrt.locking(fh.fileno(), msvcrt.LK_UNLCK, 1)
    except OSError:
        pass


def _append_to_queue(raw_args):
    """把原始路径参数(可能是文件夹/文件) append 到 queue, 由 leader 统一展开处理."""
    fh = _open_lockfile(QUEUE_LOCK)
    try:
        if not _acquire_blocking(fh):
            pass
        try:
            with open(QUEUE_FILE, 'a', encoding='utf-8') as qf:
                for p in raw_args:
                    qf.write(str(p) + '\n')
        finally:
            _release(fh)
    finally:
        fh.close()


def _drain_queue():
    """读出 queue 全部行并清空文件. 返回原始路径字符串列表."""
    fh = _open_lockfile(QUEUE_LOCK)
    try:
        if not _acquire_blocking(fh):
            return []
        try:
            if not QUEUE_FILE.exists():
                return []
            with open(QUEUE_FILE, 'r', encoding='utf-8') as qf:
                lines = qf.read().splitlines()
            with open(QUEUE_FILE, 'w', encoding='utf-8') as qf:
                pass
            return [ln for ln in lines if ln.strip()]
        finally:
            _release(fh)
    finally:
        fh.close()


def _show_summary_window(results, xlsx_path=None):
    """results: list of (final_path, status, message). 弹一个 Tk 汇总窗口."""
    try:
        import tkinter as tk
        from tkinter import scrolledtext
    except Exception:
        return

    counts = {'renamed': 0, 'skipped': 0, 'failed': 0}
    for _, status, _ in results:
        counts[status] = counts.get(status, 0) + 1

    root = tk.Tk()
    root.title('发票重命名结果')
    root.geometry('760x440')

    icon_path = SCRIPT_DIR / 'assets' / 'icon.ico'
    if icon_path.exists():
        try:
            root.iconbitmap(str(icon_path))
        except Exception:
            pass

    header_text = (
        f'重命名: {counts["renamed"]}    '
        f'跳过: {counts["skipped"]}    '
        f'失败: {counts["failed"]}    '
        f'(总 {len(results)})'
    )
    header = tk.Label(root, text=header_text,
                      font=('Microsoft YaHei', 12, 'bold'), pady=8)
    header.pack(fill='x')

    txt = scrolledtext.ScrolledText(root, font=('Consolas', 10), wrap='none')
    txt.pack(fill='both', expand=True, padx=10, pady=4)
    txt.tag_config('ok', foreground='#1b8a3a')
    txt.tag_config('skip', foreground='#777777')
    txt.tag_config('fail', foreground='#c62828')

    if not results:
        txt.insert('end', '(没有任何 PDF 被处理)\n', 'skip')
    for pdf, status, msg in results:
        if status == 'renamed':
            txt.insert('end', f'[ OK ] {pdf.name}  ->  {msg}\n', 'ok')
        elif status == 'skipped':
            txt.insert('end', f'[SKIP] {pdf.name}  ({msg})\n', 'skip')
        else:
            txt.insert('end', f'[FAIL] {pdf.name}  -- {msg}\n', 'fail')

    if xlsx_path:
        txt.insert('end', f'\n[XLSX] 已生成: {xlsx_path}\n', 'ok')
    txt.config(state='disabled')

    btn = tk.Button(root, text='关闭', command=root.destroy, width=12,
                    font=('Microsoft YaHei', 10))
    btn.pack(pady=8)

    root.attributes('-topmost', True)
    root.update()
    root.attributes('-topmost', False)
    root.mainloop()


# ---------------------------------------------------------------------------
# 入口: silent / direct 两条路径
# ---------------------------------------------------------------------------

def silent_main(args, show_summary, want_xlsx):
    """静默 + leader 路径: 写队列, 抢锁, 抢到的当 leader 处理全部."""
    if not args:
        args = [os.getcwd()]
    _append_to_queue(args)

    leader_fh = _open_lockfile(LEADER_LOCK)
    try:
        if not _acquire_nonblocking(leader_fh):
            return
        try:
            time.sleep(DEBOUNCE_SECONDS)

            collected_args = list(args)  # 用于 xlsx 目标目录解析
            all_results = []  # (final_path, status, message)
            xlsx_rows = []
            all_pdfs_processed = []

            while True:
                raw_paths = _drain_queue()
                if not raw_paths:
                    break
                for raw in raw_paths:
                    if raw not in collected_args:
                        collected_args.append(raw)
                pdfs = []
                seen = set()
                for raw in raw_paths:
                    p = Path(raw)
                    if not p.exists():
                        log_line(f'FAIL  (silent) 路径不存在: {raw}')
                        continue
                    for pdf in collect_pdfs(p):
                        key = str(pdf.resolve())
                        if key in seen:
                            continue
                        seen.add(key)
                        pdfs.append(pdf)
                for pdf in pdfs:
                    status, msg, meta, final = process_pdf(pdf)
                    if status == 'failed':
                        log_line(f'FAIL  {pdf.name}  原因: {msg}')
                    elif status == 'skipped':
                        log_line(f'SKIP  {pdf.name}  ({msg})')
                    all_results.append((final, status, msg))
                    all_pdfs_processed.append(final)
                    if status in ('renamed', 'skipped') and meta.get('amount'):
                        xlsx_rows.append(_row_from_result(status, msg, meta, final))

            xlsx_written = None
            if want_xlsx and xlsx_rows:
                target_dir = _resolve_xlsx_target_dir(collected_args, all_pdfs_processed)
                xlsx_path = _xlsx_output_path(target_dir)
                if write_summary_xlsx(xlsx_rows, xlsx_path):
                    log_line(f'XLSX  导出 -> {xlsx_path}  ({len(xlsx_rows)} 行)')
                    xlsx_written = xlsx_path

            if show_summary:
                _show_summary_window(all_results, xlsx_path=xlsx_written)
        finally:
            _release(leader_fh)
    finally:
        leader_fh.close()


def direct_main(args, want_xlsx):
    if not args:
        args = [os.getcwd()]

    targets = []
    for arg in args:
        p = Path(arg)
        if not p.exists():
            print(c(f'[ERROR] 路径不存在: {arg}', 'red'))
            continue
        targets.extend(collect_pdfs(p))

    if not targets:
        print(c('[INFO] 没有找到 PDF 文件', 'yellow'))
        return

    print(c(f'\n=== 发票重命名 (共 {len(targets)} 个 PDF) ===\n', 'cyan'))

    counts = {'renamed': 0, 'skipped': 0, 'failed': 0}
    failures = []
    xlsx_rows = []
    final_paths = []

    for pdf in targets:
        status, msg, meta, final = process_pdf(pdf)
        counts[status] += 1
        final_paths.append(final)
        if status == 'renamed':
            print(c('[ OK ]', 'green'),
                  f'{pdf.name}',
                  c('->', 'gray'),
                  c(msg, 'cyan'))
        elif status == 'skipped':
            print(c('[SKIP]', 'gray'),
                  f'{pdf.name}',
                  c(f'({msg})', 'gray'))
        else:
            print(c('[FAIL]', 'red'),
                  f'{pdf.name}',
                  c(f'-- {msg}', 'red'))
            failures.append((pdf.name, msg))
            log_line(f'FAIL  {pdf.name}  原因: {msg}')
        if status in ('renamed', 'skipped') and meta.get('amount'):
            xlsx_rows.append(_row_from_result(status, msg, meta, final))

    print()
    print(c(
        f'重命名: {counts["renamed"]}  跳过: {counts["skipped"]}  失败: {counts["failed"]}',
        'cyan'))

    if failures:
        print(c('\n以下文件需手动处理:', 'red'))
        for n, m in failures:
            print(f'  - {n}')
            print(f'      原因: {m}')

    if want_xlsx and xlsx_rows:
        target_dir = _resolve_xlsx_target_dir(args, final_paths)
        xlsx_path = _xlsx_output_path(target_dir)
        if write_summary_xlsx(xlsx_rows, xlsx_path):
            log_line(f'XLSX  导出 -> {xlsx_path}  ({len(xlsx_rows)} 行)')
            print(c(f'\n[XLSX] 已生成: {xlsx_path}', 'cyan'))


def main():
    raw = sys.argv[1:]
    silent = False
    show_summary = False
    want_xlsx = False
    paths = []
    for a in raw:
        if a == '--silent':
            silent = True
        elif a == '--summary':
            show_summary = True
        elif a == '--xlsx':
            want_xlsx = True
        else:
            paths.append(a)

    if silent:
        silent_main(paths, show_summary, want_xlsx)
    else:
        direct_main(paths, want_xlsx)


if __name__ == '__main__':
    main()
    try:
        if os.environ.get('RENAME_INVOICE_PAUSE') == '1':
            input('\n按回车键退出...')
    except (EOFError, KeyboardInterrupt):
        pass
